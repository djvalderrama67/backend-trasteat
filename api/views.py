from django.http import JsonResponse, HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action  # Agrega esta línea para importar action
from .models import *
from .serializer import *
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from rest_framework import status
import json
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import io
from django.http import FileResponse
from openpyxl.styles.borders import Border
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink



@api_view(["GET"])
def apiOverview(request):
    api_urls = {
        "Categoria": "/categoria/",
        "Vehiculo": "/vehiculo/",
        "Bodega": "/bodega/",
        "Objeto": "/objeto/",
        "Objetos por categoria": "/categoria/<str:nombre_categoria>/",
        "Calculo": "/calculo/",
        "Generar Excel": "/generar_excel/",
        "Descargar Excel": "/descargar_excel/",
    }
    return Response(api_urls)


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer

    @action(detail=True, methods=["get"])
    def objetos_por_categoria(
        self, request, nombre_categoria=None
    ):  # Recibe el parámetro nombre_categoria
        try:
            categoria = Categoria.objects.get(
                nombre=nombre_categoria
            )  # Obtén la categoría por su nombre
            objetos = Objeto.objects.filter(categoria=categoria)
            serializer = ObjetoSerializer(objetos, many=True)
            return Response(serializer.data)
        except Categoria.DoesNotExist:
            return Response(
                {"message": "Categoría no encontrada"}, status=status.HTTP_404_NOT_FOUND
            )


@csrf_exempt
@api_view(["POST"])
def generar_excel(request):
    libro_excel = openpyxl.Workbook()
    hoja_excel = libro_excel.active
    if request.method == "POST":
        try:
            # Obtener datos del cuerpo de la solicitud
            data = json.loads(request.body.decode("utf-8"))

            # Verificar que "data" está presente y es una lista
            if "data" not in data or not isinstance(data["data"], list):
                return JsonResponse(
                    {"error": "Formato de datos incorrecto"}, status=400
                )
            #img_path = "./images/Logo.jpg"
            #img = Image(img_path)

            #img.width = 1200
            #img.height = 300

            # Añadir la imagen a la hoja de cálculo
            #hoja_excel.add_image(img, 'J1')
            # Sección: Cantidad total de artículos y Volumen en m3
            cantidad_total = sum(item.get("cantidad", 0) for item in data["data"])
            volumen_total = sum(item.get("volumen", 0) for item in data["data"])
            nombre_vehiculo = data['data'][2]['vehiculo']['nombre']
            capacidad_min = data['data'][2]['vehiculo']['capacidad_min']
            capacidad_max = data['data'][2]['vehiculo']['capacidad_max']
            nombre_bodega = data['data'][3]['bodega']['nombre']
            area_bodega = data['data'][3]['bodega']['area']
            volumen_bodega = data['data'][3]['bodega']['volumen']
            objetos_seleccionados = data['data'][4].get('objetos', [])
            
            # Ajustar el ancho de la columna A a 5.14
            hoja_excel.column_dimensions['A'].width = 5.14
            hoja_excel.column_dimensions['B'].width = 13.71
            hoja_excel.column_dimensions['C'].width = 25.29
            hoja_excel.column_dimensions['D'].width = 5.29
            hoja_excel.column_dimensions['E'].width = 34.86
            hoja_excel.column_dimensions['F'].width = 8
            hoja_excel.column_dimensions['I'].width = 11
            hoja_excel.column_dimensions['H'].width = 11
            hoja_excel.column_dimensions['G'].width = 7

            
            hoja_excel.row_dimensions[1].height = 24
            hoja_excel.row_dimensions[2].height = 9.5
            hoja_excel.row_dimensions[3].height = 50
            hoja_excel.row_dimensions[4].height = 9.5
            hoja_excel.row_dimensions[5].height = 27
            hoja_excel.row_dimensions[6].height = 27
            hoja_excel.row_dimensions[7].height = 27
            hoja_excel.row_dimensions[8].height = 20
            hoja_excel.row_dimensions[9].height = 2


            # Agregar encabezados y valores
            hoja_excel['B1'] = 'Mi inventario en TrasteaT'
            hoja_excel['B3'] = 'Cantidad total de artículos:'
            hoja_excel['D3'] = cantidad_total
            hoja_excel['E3'] = 'Volumen Total:'
            hoja_excel['F3'] = volumen_total

            hoja_excel['B5'] = 'Vehículo Requerido:'
            hoja_excel['D5'] = nombre_vehiculo
            hoja_excel['F5'] = capacidad_min
            hoja_excel['G5'] = '-'
            hoja_excel['H5'] = capacidad_max
            hoja_excel['I5'] = 'T'
            
            hoja_excel['B6'] = 'Bodega Requerida:'
            hoja_excel['D6'] = nombre_bodega
            hoja_excel['F6'] = 'Area:'
            hoja_excel['G6'] = area_bodega
            hoja_excel['H6'] = 'Volumen:'
            hoja_excel['I6'] = volumen_bodega

            hoja_excel['B7'] = 'Inventario'
            hoja_excel['B8'] = 'Cantidad'
            hoja_excel['D8'] = 'Objetos'

            # Agregar datos del inventario a la hoja
            for i, objeto in enumerate(objetos_seleccionados, start=10):
                nombre_objeto = objeto.get('nombre', '')
                cantidad_objeto = objeto.get('cantidad', 0)

                hoja_excel[f'D{i}'] = nombre_objeto
                hoja_excel[f'B{i}'] = cantidad_objeto

            font_format = Font(size=14, name='Comic Sans MS', color='F9C307', bold=True)



            # Combinar celdas para el título
            hoja_excel.merge_cells('B1:I1')
            hoja_excel.merge_cells('B7:I7')
            hoja_excel.merge_cells('B3:C3')
            hoja_excel.merge_cells('B5:C5')
            hoja_excel.merge_cells('D5:E5')
            hoja_excel.merge_cells('B6:C6')
            hoja_excel.merge_cells('D6:E6')
            hoja_excel.merge_cells('B8:C8')
            hoja_excel.merge_cells('D8:I8')
            hoja_excel.merge_cells('F8:I8')
            hoja_excel.merge_cells('F3:I3')
            ultima_fila = hoja_excel.max_row

            # Iterar a través de las filas desde la fila 9 hasta la última fila en la columna I
            for fila in range(9, ultima_fila + 1):
                hoja_excel.merge_cells(f'D{fila}:I{fila}')
            hoja_excel['B1'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B7'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B5'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B6'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['B8'].alignment = Alignment(horizontal='center', vertical='center')
            hoja_excel['D8'].alignment = Alignment(horizontal='center', vertical='center')
            for row_number in range(8, 110):
                hoja_excel.merge_cells(f'B{row_number}:C{row_number}')
                hoja_excel.merge_cells(f'D{row_number}:E{row_number}')
                hoja_excel.merge_cells(f'F{row_number}:G{row_number}')
                
                hoja_excel[f'B{row_number}'].alignment = Alignment(horizontal='center', vertical='center')
                hoja_excel[f'D{row_number}'].alignment = Alignment(horizontal='center', vertical='center')
                hoja_excel[f'F{row_number}'].alignment = Alignment(horizontal='center', vertical='center')

            hoja_excel['B1'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['B3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['C3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['D3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['E3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['F3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['G3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['H3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['I3'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['B6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['C6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['D6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['E6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['F6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['G6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['H6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['I6'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['B5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['C5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['D5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['E5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['F5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['G5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['H5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['I5'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['B7'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['B8'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['C8'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['D8'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['E8'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['F8'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            hoja_excel['G8'].fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            # Establecer el tamaño y el tipo de letra para toda la hoja
            color_letra = 'F9C307'
            for row in hoja_excel.iter_rows(min_row=1, max_col=hoja_excel.max_column, max_row=8):
                for cell in row:
                    cell.font = Font(size=14, name='Comic Sans MS', color=color_letra,)
            hoja_excel = libro_excel.active


            # Define el color de celda y color de fuente
            color_celda = '1A1A1A'
            color_fuente = 'F9C307'

            # Define el estilo de fuente para negrita, Comic Sans MS, tamaño 14 y color de fuente
            font_bold = Font(size=14, name='Comic Sans MS', bold=True, color=color_fuente)

            # Define el color de relleno para las celdas
            fill_color = PatternFill(start_color=color_celda, end_color=color_celda, fill_type='solid')

            # Lista de celdas para aplicar el formato
            celdas_estilo = ['B1', 'B3', 'E3', 'B5', 'B6', 'F6', 'H6', 'B7', 'B8', 'D8']

            # Aplica el estilo a las celdas
            for celda in celdas_estilo:
                hoja_excel[celda].font = font_bold
                hoja_excel[celda].fill = fill_color


            # Establecer el rango de celdas desde A1 hasta la última fila que desees (por ejemplo, A100)
            rango_celdas = hoja_excel['A1:I100']  # Modifica el rango según tus necesidades

            # Aplicar la alineación centrada a todas las celdas en el rango especificado
            for row in rango_celdas:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            rango_bordes = hoja_excel['B1:I8']

            # Crear un borde blanco
            borde_blanco = Border(left=Side(border_style='thin', color='FFFFFF'),
                                right=Side(border_style='thin', color='FFFFFF'),
                                top=Side(border_style='thin', color='FFFFFF'),
                                bottom=Side(border_style='thin', color='FFFFFF'))
            
            # Identificar la última fila en la columna I (aquí asumimos que la columna I contiene datos)
            ultima_fila = hoja_excel.max_row

            celdas_negrita = ['B1', 'B3', 'E3', 'B5', 'B6', 'F6', 'H6', 'B7', 'B8', 'E8']



            # Establecer el formato para las celdas desde B9 hasta la última fila en la columna I
            for fila in range(9, ultima_fila + 1):
                for columna in range(2, 10):  # Columna B a I
                    celda = hoja_excel.cell(row=fila, column=columna)
                    celda.font = Font(size=14, name='Comic Sans MS', color='1A1A1A')

            # Identificar la última fila en la columna I (aquí asumimos que la columna I contiene datos)
            ultima_fila = hoja_excel.max_row

            # Establecer el formato para las celdas desde B9 hasta la última fila en la columna I
            for fila in range(9, ultima_fila + 1):
                for columna in range(2, 10):  # Columna B a I
                    celda = hoja_excel.cell(row=fila, column=columna)
                    if fila % 2 == 0:  # Alternar entre dos colores en filas pares e impares
                        celda.fill = PatternFill(start_color='C4BD97', end_color='C4BD97', fill_type='solid')
                    else:
                        celda.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            for fila in range(9, ultima_fila + 1):
                for columna in range(2, 10):  # Columna B a I
                    celda = hoja_excel.cell(row=fila, column=columna)
                    celda.border = borde_blanco

            # Aplicar el borde blanco al rango de celdas especificado
            for row in rango_bordes:
                for cell in row:
                    cell.border = borde_blanco

            # for objeto in data['data'][4]['objetos']:
            #     nombre_objeto = objeto['nombre']
            #     cantidad_objeto = objeto['cantidad']
            #     volumen_objeto = objeto['volumen']
            # print("Nombre del objeto:", nombre_objeto)
            # print("Cantidad del objeto:", cantidad_objeto)
            # print("Volumen del objeto:", volumen_objeto)
            

            excel_buffer = io.BytesIO()
            libro_excel.save(excel_buffer)

            # Devolver el archivo Excel como respuesta
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f"attachment; filename=inventario.xlsx"

            # Guardar el archivo en el sistema de archivos si es necesario
            excel_file_path = os.path.join(settings.MEDIA_ROOT, f"inventario.xlsx")
            with open(excel_file_path, "wb") as file:
                file.write(excel_buffer.getvalue())

            # Devolver una respuesta exitosa
            return response

        except Exception as e:
            # Manejar errores
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)


def descargar_excel(request):
    excel_file_path = os.path.join(settings.MEDIA_ROOT, "inventario.xlsx")

    if os.path.exists(excel_file_path):
        with open(excel_file_path, "rb") as file:
            response = HttpResponse(
                file.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = "attachment; filename=inventario.xlsx"

        return response
    else:
        return JsonResponse({"error": "El archivo Excel no existe"}, status=404)


class VehiculoViewSet(viewsets.ModelViewSet):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer


class BodegaViewSet(viewsets.ModelViewSet):
    queryset = Bodega.objects.all()
    serializer_class = BodegaSerializer


class ObjetoViewSet(viewsets.ModelViewSet):
    queryset = Objeto.objects.all()
    serializer_class = ObjetoSerializer


class CalculoViewSet(viewsets.ModelViewSet):
    queryset = Calculo.objects.all()
    serializer_class = CalculoSerializer
